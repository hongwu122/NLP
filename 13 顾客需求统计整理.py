# -*- coding:utf-8 -*-
import openpyxl
import pandas as pd
from collections import Counter
from openpyxl.styles import Color, Font, Alignment, Border, Side, PatternFill, colors


def get_max_row(sheet): # 获取最大行
    i=sheet.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i-1
        else:
            real_max_row = i
            break

    return real_max_row


color = PatternFill("solid", fgColor="00FFFF99")
font = Font(name=u'宋体', bold=True, italic=False, size=11)  # bold是否加粗, italic是否斜体

def main(path, to_excel):
    workbook = openpyxl.load_workbook(path)
    sheetnames = workbook.sheetnames # ['售后服务', '外观简洁好看', '显示效果良好', '机身轻巧纤薄', '电池耐用度高', '运行流畅', '人机界面友好', '零配件齐全', '网络信号稳定']
    print(workbook.sheetnames) # ['速度', '处理器', 'CPU', '骁龙', '四核', '单核', 'GPU', '内存', '容量', 'RAM', 'ROM', '运行速度', '反应速度', '延迟', '存储', '空间', 'SD', '系统', '操作', '流畅', 'IOS', 'Android', '黑屏', 'Bug', '兼容', '刷机', '闪退', 'root', '死机', '启动', '卡顿', '卡机', '不卡', '越狱', '应用', '程序', 'APP', '功能', '软件', '导航', '网页', '浏览器', '地图', '娱乐', '玩游戏', '手游', '王者', '', '信息', 'GPS', '开机']

    df_Feature = pd.read_excel('工具/Dictionary/电脑产品特征词.xlsx')
    dicts = {}
    feature_word_range_list = []
    for d in range(len(df_Feature['顾客需求'])):
        customer_demand = df_Feature['顾客需求'][d]
        feature_word = str(df_Feature['特征词'][d]).split(' ')
        # print(feature_word) # ['屏幕', '桌面', '触摸屏', '电容屏', '显示屏', '主屏', '弧面', '触屏', '曲屏', '分辨率', '亮度', '显示', '蓝光', '界面', '色彩', '画质', '清晰度', '画面', '刘海', '内屏', '分屏', '饱和度', '色泽']
        # print(len(feature_word)) # 23
        feature_word_range_list.append(len(list(set([f for f in feature_word if f != ''])))) # 要去重
        dicts.update({customer_demand : list(set([f for f in feature_word if f != '']))})  # feature_word =  [f for f in feature_word if f != '']
    print(dicts) # {'显示效果良好': ['屏幕', '桌面', '触摸屏', '电容屏', '显示屏', '主屏', '弧面', '触屏', '曲屏', '分辨率', '亮度', '显示', '蓝光', '界面', '色彩', '画质', '清晰度', '画面', '刘海', '内屏', '分屏', '饱和度', '色泽'], '网络信号稳定': ['网络', '信号', '上网', 'WLAN', '4G', '蓝牙', '双频', 'WIFI', '红外', '无线', '网速'], '电池耐用度高': ['电池', '待机', '充电器', '充电', '快充', '电量', '耗电', '用电', '时长', '续航', '发热', '发烧', '毫安', '电源', '省电', '发烫', '散热'], '外观简洁好看': ['外观', '外形', '机身', '体积', '缝隙', '工艺', '颜色', '外表', '线条', '机身', '样子', '造型', '设计', '个性', '外壳', '', '尺寸', '机型', '颜值', '美观', '漂亮', '划痕', '厚度', '边框', '后盖', '后壳', '大小', '耐看'], '运行流畅': ['速度', '处理器', 'CPU', '骁龙', '四核', '单核', 'GPU', '内存', '容量', 'RAM', 'ROM', '运行速度', '反应速度', '延迟', '存储', '空间', 'SD', '系统', '操作', '流畅', 'IOS', 'Android', '黑屏', 'Bug', '兼容', '刷机', '闪退', 'root', '死机', '启动', '卡顿', '卡机', '不卡', '越狱', '应用', '程序', 'APP', '功能', '软件', '导航', '网页', '浏览器', '地图', '娱乐', '玩游戏', '手游', '王者', '', '信息', 'GPS', '开机'], '零配件齐全': ['配件', '壳', '膜', '玻璃膜', '保护膜', '套', '手机套', '耳机', '数据线', '电源适配器'], '机身轻巧纤薄': ['手感', '重量', '质感', '触感', '材质', '做工', '材料', '薄', '轻'], '人机界面友好': ['指纹', '解锁', '人脸识别', '语音识别', '多屏协同', '移动应用引擎', 'HMS', '电视剧', '音箱', '听歌', '爆音', '杂音', '听筒', '通话', '噪音', '音效', '闹铃', '摄像'], '售后服务': ['售后', '服务', '物流', '客服', '态度', '口碑', '顺丰', '卖家', '维修', '保修', '质保', '包装', '送货']}
    print(feature_word_range_list) # [23, 11, 17, 28, 51, 10, 9, 18, 13]

    result_workbook = openpyxl.Workbook()
    result_worksheet = result_workbook.worksheets[0]
    # 写入表头
    result_worksheet.cell(1, 1, '序号').font = font
    result_worksheet.cell(1, 2, '顾客需求').font = font
    result_worksheet.cell(1, 3, '提到此顾客需求的评论条数').font = font
    for n in range(max(feature_word_range_list)):
        result_worksheet.cell(1, (2+n)*2, '特征词{}'.format(n+1)).font = font
        result_worksheet.cell(1, (2+n)*2, '特征词{}'.format(n+1)).fill = color
        result_worksheet.cell(1, (2+n)*2+1, '数量').font = font

    for i in range(len(workbook.worksheets)):
        worksheet = workbook.worksheets[i]
        xuhao = i + 1
        sheetname = sheetnames[i]
        all_num = get_max_row(worksheet) - 1

        df = pd.read_excel(path, sheet_name=sheetname)
        feature_words = df['包含产品特征词'].values.tolist() # .split(' ')
        if '' in feature_words:
            feature_words.remove('')
        print(feature_words)

        new_feature_words = []
        for f in feature_words:
            new = f.split(' ')
            for n in new:
                new_feature_words.append(n)
        print(new_feature_words)

        result_dict = Counter(new_feature_words)
        print(result_dict)

        result_worksheet.cell(2+i, 1, str(xuhao))
        result_worksheet.cell(2+i, 2, str(sheetname))
        result_worksheet.cell(2+i, 3, str(all_num))

        keys_list = list(result_dict.keys())
        values_list = list(result_dict.values())
        print(dicts[sheetname])
        print(keys_list)
        print(values_list)

        # word_ = dicts[sheetname]
        # if '' in word_:
        #     word_.remove('')
        word_no_count = [word for word in dicts[sheetname] if word not in keys_list and word != '']
        print('#############\n',word_no_count)

        for n in range(len(keys_list)):
            result_worksheet.cell(2+i, (2+n)*2, str(keys_list[n])).font = font
            result_worksheet.cell(2+i, (2+n)*2, str(keys_list[n])).fill = color
            result_worksheet.cell(2+i, (2+n)*2+1, str(values_list[n]))

        if word_no_count != []:
            for n in range(len(word_no_count)):
                result_worksheet.cell(2+i, (2+n + len(keys_list))*2, str(word_no_count[n])).font = font
                result_worksheet.cell(2+i, (2+n + len(keys_list))*2, str(word_no_count[n])).fill = color
                result_worksheet.cell(2+i, (2+n + len(keys_list))*2+1, 0)

    result_workbook.save(to_excel)

# 显示效果良好
# 网络信号稳定
# 电池耐用度高
# 外观简洁好看
# 运行流畅
# 零配件齐全
# 机身轻巧纤薄
# 人机界面友好
# 售后服务

# '工具/Dictionary/电脑产品特征词.xlsx'
# 屏幕 桌面 触摸屏 电容屏 显示屏 主屏 弧面 触屏 曲屏 分辨率 亮度 显示 蓝光 界面 色彩 画质 清晰度 画面 刘海 内屏 分屏 饱和度 色泽
# 网络 信号 上网 WLAN 4G 蓝牙 双频 WIFI 红外 无线 网速
# 电池 待机 充电器 充电 快充 电量 耗电 用电 时长 续航 发热 发烧 毫安 电源 省电 发烫 散热
# 外观 外形 机身 体积 缝隙 工艺 颜色 外表 线条 机身 样子 造型 设计 个性 外壳  尺寸 机型 颜值 美观 漂亮 划痕 厚度 边框 后盖 后壳 大小 耐看
# 速度 处理器 CPU 骁龙 四核 单核 GPU 内存 容量 RAM ROM 运行速度 反应速度 延迟 存储 空间 SD 系统 操作 流畅 IOS Android 黑屏 Bug 兼容 刷机 闪退 root 死机 启动 卡顿 卡机 不卡 越狱 应用 程序 APP 功能 软件 导航 网页 浏览器 地图 娱乐 玩游戏 手游 王者  信息 GPS 开机
# 配件 壳 膜 玻璃膜 保护膜 套 手机套 耳机 数据线 电源适配器
# 手感 重量 质感 触感 材质 做工 材料 薄 轻
# 指纹 解锁 人脸识别 语音识别 多屏协同 移动应用引擎 HMS 电视剧 音箱 听歌 爆音 杂音 听筒 通话 噪音 音效 闹铃 摄像
# 售后 服务 物流 客服 态度 口碑 顺丰 卖家 维修 保修 质保 包装 送货



if __name__ == '__main__':
    # main('数据/19 按顾客需求分类整理2.xlsx','数据/20 顾客需求统计整理2.xlsx')

    main('数据/19 按顾客需求分类整理2.2.xlsx','数据/20 顾客需求统计整理2.2.xlsx')










