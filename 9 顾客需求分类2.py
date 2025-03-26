# 序号	顾客需求
# 1	显示效果良好
# 2	网络信号稳定
# 3	电池耐用度高
# 4	外观简洁好看
# 5	运行流畅
# 6	零配件齐全
# 7	机身轻巧纤薄
# 8	人机界面友好
# 9	售后服务
import pandas as pd
import openpyxl
import os, re

def main(dfData, to_excel):
    demand_word = ['显示效果良好', '网络信号稳定', '电池耐用度高', '外观简洁好看', '运行流畅', '零配件齐全', '机身轻巧纤薄', '人机界面友好', '售后服务']
    df = pd.DataFrame(columns=("对应评论句子", "包含产品特征词", "对应顾客需求", "句子长度", "对应评论星级", "情感预测", "情感计算值"))
    for index, row in dfData.iterrows():
        customer_demand = list(str(dfData.loc[index, '所属顾客需求']).split(' '))
        production_sentence = eval(dfData.loc[index, '提到产品特征的句子'])
        production_feature = list(str(dfData.loc[index, '含有产品特征词']).split(' '))
        xingji = dfData.loc[index, '星级']
        if '' in customer_demand:
            customer_demand.remove('')
        if '' in production_feature:
            production_feature.remove('')


        # print(customer_demand) # ['运行流畅', '显示效果良好', '显示效果良好', '电池耐用度高', '外观简洁好看', '机身轻巧纤薄', '机身轻巧纤薄', '售后服务']
        # print(len(customer_demand),len(production_sentence))
        for i in range(len(customer_demand)):
            cd = customer_demand[i]
            if cd in demand_word:
                sentence = production_sentence[i]
                feature = production_feature[i]
                words_num = len(re.findall(r'[\u4E00-\u9FFF]', sentence))  # 字数
                dict = {"对应评论句子": sentence,
                 "包含产品特征词": feature,
                 "对应顾客需求":cd,
                 "句子长度": words_num,
                 "对应评论星级":xingji,
                 "情感预测": None,
                 "情感计算值": None,
                 }

                print(cd)
                print(dict)
                df = df.append(dict, ignore_index=True)

    df.to_excel(to_excel)

def open(to_excel):
    workbook = openpyxl.Workbook()
    demand_word = ['显示效果良好', '网络信号稳定', '电池耐用度高', '外观简洁好看', '运行流畅', '零配件齐全', '机身轻巧纤薄', '人机界面友好', '售后服务']
    for dw in demand_word:
        worksheet = workbook.create_sheet(dw, demand_word.index(dw))
        write = ['序号','对应评论句子','包含产品特征词','句子长度','情感预测','情感计算值']
        for i in range(len(write)):
            worksheet.cell(1, i+1, write[i])
    workbook.save(to_excel)

if __name__ == '__main__':
    # open(to_excel='数据/15 顾客需求分类.xlsx')
    # 获取数据
    dfData = pd.read_excel('数据/6 完整数据集（句子产品特征词性统计后）(新2）.xlsx',index_col=0)
    path = '数据/16 顾客需求分类2.xlsx'
    main(dfData, to_excel=path)