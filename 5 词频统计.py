# 论文-jieba词频分析——评论词频统计
# txt输入，xlsx输出
import jieba
import jieba.analyse
import openpyxl  # 写入Excel表的库
import pandas as pd

def worldCount(dfData):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    word_lst = []
    key_list = []
    for line in dfData['分词后']:
        item = line.strip('\n\r').split('\t')  # 制表格切分
        tags = jieba.analyse.extract_tags(item[0])  # jieba分词
        for t in tags:
            word_lst.append(t)

    word_dict = {}
    with open(r'数据/8 词频统计.txt', 'w+', encoding='utf-8') as wf2:  # 打开文件

        for item in word_lst:
            if item not in word_dict:  # 统计数量
                word_dict[item] = 1
            else:
                word_dict[item] += 1

        orderList = list(word_dict.values())
        orderList.sort(reverse=True)
        # print orderList
        for i in range(len(orderList)):
            for key in word_dict:
                if word_dict[key] == orderList[i]:
                    wf2.write(key + ' ' + str(word_dict[key]) + '\n')  # 写入txt文档
                    key_list.append(key)
                    word_dict[key] = 0
    print(key_list)
    for i in range(len(key_list)):
        sheet.cell(i+1, 2, orderList[i])
        sheet.cell(i+1, 1, key_list[i])
    workbook.save('8 词频统计.xlsx')  # 保存为 wordCount.xls文件


if __name__ == "__main__":
    dfData = pd.read_excel('数据/7 完整数据集（情感词典预测）.xlsx',index_col=0)
    worldCount(dfData)
