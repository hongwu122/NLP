
from selenium import webdriver
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support.wait import WebDriverWait
from pyquery import PyQuery as pq
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
import time

# MAX_PAGE = 10000 # 最大页数
driver_path = r"C:\chromedriver\chromedriver108.0.5359.71.exe"  # 调用软件地址
browser = webdriver.Chrome(executable_path=driver_path)
wait = WebDriverWait(browser, 10)
# 最大化窗口
browser.maximize_window()

workbook = openpyxl.Workbook()
worksheet = workbook.active

biaotou = ['评论者', '会员', '星级', '评论内容', '大小', '版本', '点赞数', '评论回复数', '发表时间']
for w in range(len(biaotou)):
    worksheet.cell(1, w + 1, biaotou[w])
count = 2


# 构造URL请求
def index_page(url,n,c):
    global count, workbook, worksheet
    try:
        browser.get(url=url)
        # 等待评论加载完成
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#comment .mc .tab-con .comment-item')))
        nn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,n)))
        nn.click()
        time.sleep(10)
        for i in range(1, 101):
            try:
                print('\n\n##################','当前第{}页'.format(i))
                # 等待评论加载完成
                wait.until(EC.presence_of_element_located( (By.CSS_SELECTOR, '#comment .mc .tab-con .comment-item')))

                get_comments()

                # 点击下一页
                next = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, c)))
                next.click()
                time.sleep(10)
            except:
                return
    except:
        return


def get_comments():
    global count
    """
    提取评价数据
    """
    html = browser.page_source  # 获取页码源码
    # print(html)
    doc = pq(html)
    items = doc('#comment .mc .tab-con .comment-item').items()  # 提取评论目录
    for item in items:

        Name = item.find('.user-column .user-info').text()   # 提取用户名，如:z***0
        level = item.find('.user-column .user-level').text()  # 提取用户名会员，如:PLUS会员
        Star = str(item.find('.comment-column.J-comment-column').children('.comment-star')).strip()[-4]  # 提取评分，如:5
        Comment = item.find('.comment-column .comment-con').text()  # 提取评价内容，如:很早就想买一个iPad 可以用来看
        addComment = item.find('.append-comment .comment-con').text()  # 提取评价内容，如:很早就想买一个iPad 可以用来看
        Size = item.find('.comment-column .comment-message .order-info span:nth-child(2)').text()  # 提取商品大小，如:WIFI版128G
        Version = item.find('.comment-column .comment-message .order-info span:nth-child(1)').text()  # 提取商品版本，如:Pencil套装版
        Time = item.find('.comment-column .comment-message .order-info span:nth-child(5)').text()  # 提取评价时间，如:2018-11-17 00:07
        zan = item.find('.comment-column .comment-message .comment-op a:nth-child(2)').text() # 点赞数
        pinlun = item.find('.comment-column .comment-message .comment-op a:nth-child(3)').text() # 评论数

        write_content = [Name,level,Star,Comment,Size,Version,zan,pinlun,Time]
        print(write_content)
        for w in range(len(write_content)):
            worksheet.cell(count, w+1, write_content[w])
        count += 1


if __name__ == '__main__':
    url_list = [
         'https://item.jd.com/100016960357.html#comment',
        'https://item.jd.com/10052006595898.html#comment', 'https://item.jd.com/10048885630241.html#comment',
        'https://item.jd.com/10048899122421.html#comment','https://item.jd.com/28355253535.html#comment',
        'https://item.jd.com/64986169835.html#comment','https://item.jd.com/100018005951.html#comment',
        'https://item.jd.com/64975124552.html#comment','https://item.jd.com/10042473909988.html#comment',
        'https://item.jd.com/100039356470.html','https://item.jd.com/10048453831960.html#comment' ,
        'https://item.jd.com/10053566761048.html#comment','https://item.jd.com/68628594385.html#comment' ,
        'https://item.jd.com/68628594387.html#comment','https://item.jd.com/100018076011.html#comment',
        'https://item.jd.com/10066859460426.html#comment','https://item.jd.com/10049199599316.html#comment' ,
        'https://item.jd.com/10049199599316.html#comment','https://item.jd.com/100023215451.html#comment' ,
        'https://item.jd.com/62542227740.html#comment', 'https://item.jd.com/100035034729.html#comment' ,
        'https://item.jd.com/62579483893.html#comment'
    ]
    url_list2 = [
    'https://item.jd.com/100029902363.html#comment',
    'https://item.jd.com/100042886045.html#comment',
    'https://item.jd.com/100029902383.html#comment',
    'https://item.jd.com/100041048682.html#comment',
    'https://item.jd.com/100033982753.html#comment',
    'https://item.jd.com/100017947815.html#comment',
    'https://item.jd.com/100034269339.html#comment',
    'https://item.jd.com/100042984679.html#comment',
    'https://item.jd.com/100042885991.html#comment',
    'https://item.jd.com/100029902359.html#comment',
    'https://item.jd.com/100018778303.html#comment',
    'https://item.jd.com/100029902389.html#comment',
    'https://item.jd.com/100042886029.html#comment',
        'https://item.jd.com/100017947833.html#comment',
        'https://item.jd.com/100039511262.html#comment',
        'https://item.jd.com/100029618081.html#comment',
        'https://item.jd.com/100022917525.html#comment',
        'https://item.jd.com/100018778303.html#comment',
        'https://item.jd.com/100016960353.html#comment',
        'https://item.jd.com/100029902339.html#comment',
        'https://item.jd.com/100032062474.html#comment',
        'https://item.jd.com/100037028281.html#comment',
        'https://item.jd.com/100039550074.html#comment',
    'https://item.jd.com/68628594389.html?extension_id=eyJhZCI6IjQ3IiwiY2giOiIyIiwic2t1IjoiNjg2Mjg1OTQzODkiLCJ0cyI6IjE2NzIwMjI2NDAiLCJ1bmlxaWQiOiJ7XCJjbGlja19pZFwiOlwiYjFkMTliMzYtOWJmOC00YWVlLWIwNTctNzYxZjM3YjQ1YWE4XCIsXCJtYXRlcmlhbF9pZFwiOlwiODA3MDk5MTg5ODgzNDM3OTMyOFwiLFwicG9zX2lkXCI6XCI0N1wiLFwic2lkXCI6XCJjZWJlMjQ5ZS1iMWU4LTQ0ZTUtYjVkNy1hYjM2NDdhYmQ0NWNcIn0ifQ==&jd_pop=b1d19b36-9bf8-4aee-b057-761f37b45aa8&abt=0#comment'
    ]
    for u in url_list2:
        print('\n\n##################','当前中评,第{}'.format(u))
        index_page(u,n='#comment > div.mc > div.J-comments-list.comments-list.ETab > div.tab-main.small > ul > li:nth-child(6)',c='#comment-5 > div.com-table-footer > div > div > a.ui-pager-next')
        print('\n\n##################','当前差评,第{}'.format(u))
        index_page(u,n='#comment > div.mc > div.J-comments-list.comments-list.ETab > div.tab-main.small > ul > li:nth-child(7)',c='#comment-6 > div.com-table-footer > div > div > a.ui-pager-next')
        print('\n\n##################','当前追评,第{}'.format(u))
        index_page(u,n='#comment > div.mc > div.J-comments-list.comments-list.ETab > div.tab-main.small > ul > li.J-addComment',c='#comment-3 > div.com-table-footer > div > div > a.ui-pager-next')
        print('\n\n##################', '当前好评,第{}'.format(u))
        index_page(u, n='#comment > div.mc > div.J-comments-list.comments-list.ETab > div.tab-main.small > ul > li:nth-child(5)', c='#comment-4 > div.com-table-footer > div > div > a.ui-pager-next')
        print('\n\n##################', '当前视频晒图,第{}'.format(u))
        index_page(u, n='# comment > div.mc > div.J-comments-list.comments-list.ETab > div.tab-main.small > ul > li:nth-child(3)', c='# comment-2 > div.com-table-footer > div > div > a.ui-pager-next')


        workbook.save('jd_hw.xlsx')



