# 序号	顾客需求
# 1	显示效果良好
# 2	网络信号稳定
# 3	电池耐用度高
# 4	外观简洁好看
# 5	运行流畅
# 6	零配件齐全
# 7	机身轻巧纤薄
# 8	人机界面友好
# 9	售后服务
import pandas as pd
import openpyxl
import os

def main(dfData, open_excel, to_excel):
    demand_word = ['显示效果良好', '网络信号稳定', '电池耐用度高', '外观简洁好看', '运行流畅', '零配件齐全', '机身轻巧纤薄', '人机界面友好', '售后服务']
    # df_list = [i for i in range(8)]

    # names = locals()
    # for i in range(8):
    #     # df = pd.DataFrame(columns=("对应评论句子", "包含产品特征词", "句子长度", "情感预测", "情感计算值"))
    #     # df_list[i] = df
    #     # exec('df{} = pd.read_excel(open_excel, sheet_name=demand_word[{}], index_col=0)'.format(i,i)) # 建立八个变量
    #     exec(f'df{i} = pd.read_excel(open_excel, sheet_name=demand_word[{i}], index_col=0)') # 建立八个变量
    #     names[f'df_{i}'] = i
        # exec('df%s = pd.DataFrame(columns=("对应评论句子", "包含产品特征词", "句子长度", "情感预测", "情感计算值"))' % i)
    df1 = pd.read_excel(open_excel, sheet_name=demand_word[0], index_col=0)
    df2 = pd.read_excel(open_excel, sheet_name=demand_word[1], index_col=0)
    df3 = pd.read_excel(open_excel, sheet_name=demand_word[2], index_col=0)
    df4 = pd.read_excel(open_excel, sheet_name=demand_word[3], index_col=0)
    df5 = pd.read_excel(open_excel, sheet_name=demand_word[4], index_col=0)
    df6 = pd.read_excel(open_excel, sheet_name=demand_word[5], index_col=0)
    df7 = pd.read_excel(open_excel, sheet_name=demand_word[6], index_col=0)
    df8 = pd.read_excel(open_excel, sheet_name=demand_word[7], index_col=0)
    df9 = pd.read_excel(open_excel, sheet_name=demand_word[8], index_col=0)

    df_list = [df1, df2, df3, df4, df5, df6, df7, df8, df9]
    # print(df_list)
    writer = pd.ExcelWriter(open_excel, engine='openpyxl')
    for index, row in dfData.iterrows():
        customer_demand = list(str(dfData.loc[index, '所属顾客需求']).split(' '))
        if '' in customer_demand:
            customer_demand.remove('')

        production_sentence = eval(dfData.loc[index, '提到产品特征的句子'])

        production_feature = list(str(dfData.loc[index, '含有产品特征词']).split(' '))
        if '' in production_feature:
            production_feature.remove('')

        print(customer_demand) # ['运行流畅', '显示效果良好', '显示效果良好', '电池耐用度高', '外观简洁好看', '机身轻巧纤薄', '机身轻巧纤薄', '售后服务']

        # print(len(customer_demand),len(production_sentence))
        for cd in customer_demand:
            if cd in demand_word:
                print(cd)
                print(customer_demand.index(cd))
                print(production_sentence)
                sentence = production_sentence[customer_demand.index(cd)]
                feature = production_feature[customer_demand.index(cd)]
                # df = pd.DataFrame(columns=("对应评论句子", "包含产品特征词", "句子长度", "情感预测", "情感计算值"))
                # df = pd.DataFrame()
                writer = pd.ExcelWriter(to_excel)
                with pd.ExcelWriter(to_excel, engine='openpyxl', mode='a') as writer:
                #     # df.to_excel(writer, sheet_name=cd)
                #     # df2.to_excel(writer, sheet_name='df2')
                #     # df[cd].to_excel(excel_writer=writer, sheet_name=cd) # , encoding="GBK"
                #     # print(df)
                # dataAdd = pd.Series(
                #     {"对应评论句子": sentence,
                #      "包含产品特征词": feature,
                #      "句子长度": len(sentence),
                #      "情感预测": None,
                #      "情感计算值": None,
                #      })
                    dict = {"对应评论句子": sentence,
                     "包含产品特征词": feature,
                     "句子长度": len(sentence),
                     "情感预测": None,
                     "情感计算值": None,
                     }
                    print(dict)
                    print(demand_word.index(cd),len(df_list))
                    # print(df_list)
                    # print(df_list[])
                    print(demand_word)
                    print(cd)
                    df = df_list[demand_word.index(cd)]
                    df.loc[len(df)] = dict
                    df.to_excel(writer, sheet_name=cd)
                    print('xxxxxxxxxxxxxxxx')
                    # print(df)
                    print('xxxxxxxxxxxxxxxx')
                # dfAdd = dfAdd.append(dataAdd, ignore_index=True)
                #
                # d_index = demand_word.index(cd)
                # print(d_index,len(df_list))
                # dfResult = pd.concat([df_list[d_index-1], dfAdd], axis=1)
                # df_list[demand_word.index(cd)-1] = dfResult
                #
                # dfResult.to_excel(writer, sheet_name=cd)
                #     # print(dfAdd)
                #     # print(dfResult)
                # dataAdd.to_excel(writer, sheet_name=cd, index=False)
        writer.save()
        writer.close()
                # for i in range(8):
                #         exec("df%s.to_excel(writer, sheet_name=demand_word[i])" % i)
                # writer.save()


    # workbook.create_sheet()
    # workbook.save(to_excel)

def open(to_excel):
    workbook = openpyxl.Workbook()
    demand_word = ['显示效果良好', '网络信号稳定', '电池耐用度高', '外观简洁好看', '运行流畅', '零配件齐全', '机身轻巧纤薄', '人机界面友好', '售后服务']
    for dw in demand_word:
        worksheet = workbook.create_sheet(dw, demand_word.index(dw))
        write = ['序号','对应评论句子','包含产品特征词','句子长度','情感预测','情感计算值']
        for i in range(len(write)):
            worksheet.cell(1, i+1, write[i])
    workbook.save(to_excel)








if __name__ == '__main__':
    open(to_excel='数据/15 顾客需求分类.xlsx')
    # 获取数据
    dfData = pd.read_excel('数据/6 完整数据集（句子产品特征词性统计后）(新2）.xlsx',index_col=0)
    #
    path = '数据/16 顾客需求分类.xlsx'
    if os.path.exists(path):
        os.remove(path)
    if not os.path.exists(path):
        print('no file path')
        workbook = openpyxl.Workbook()
        # workbook.remove(workbook.active)
        workbook.save(path)
    main(dfData, open_excel='数据/15 顾客需求分类.xlsx', to_excel=path)